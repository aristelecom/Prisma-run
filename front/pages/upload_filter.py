'''To create filters to process datasets

Receives:
raw excel dataset

Returns:
cleaned and processed tables
'''

# Imports
from functions import filter_dataframe
from pathlib import Path
import requests
import openpyxl
import boto3
import pandas as pd
import streamlit as st
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '../../'))
from libs.config import settings

st.set_page_config(
    page_title="Forever filters",
    page_icon="🚲",
    layout="wide",
)

#set up containers to work in 
header = st.container()

uploaded_xlxs_file = ""
sheetnames = []

# title
with header:
    st.title('🚲 Bikes forever store')

# Selector file with 200mb max size
uploaded_xlxs_file = st.file_uploader(
        "Upload your xlxs file into local storage", type=["xlsx"],
            accept_multiple_files=False,  
            help="To activate 'wide mode', go to the hamburger menu > Settings > turn on 'wide mode'"
)

# Make request only if file had been upload
if uploaded_xlxs_file:
    # Request to api endpoint
    response = requests.post(
        'http://127.0.0.1:8000/dataset/upload',
        data={"type": "multipart/form-data"},
        files={"file": uploaded_xlxs_file})

    # transform json object into dictionary
    response = response.json()
    print(response)
    # open and extract xlxs cheetsheet
    xlsFile = openpyxl.load_workbook(response['local_path'])
    sheetnames = xlsFile.sheetnames

    st.text("Look into upload file")
    file_container = st.expander("Check your uploaded files")
    # Import datasets from all excel sheets
    for name in sheetnames:
        file_container.write(pd.read_excel(uploaded_xlxs_file, sheet_name=name, skiprows=[0]))

with st.sidebar:
    name_option = st.selectbox('Select Table that you would like to see', sheetnames)

if name_option in sheetnames:
    df_filtered = filter_dataframe(pd.read_excel(uploaded_xlxs_file, sheet_name=name_option,skiprows=[0]))
    st.dataframe(df_filtered)
    st.download_button('Download file', data=df_filtered.to_csv(),file_name=f'{name_option}_filtered.csv')

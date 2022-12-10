'''To create filters to process datasets

Receives:
raw excel dataset

Returns:
cleaned and processed tables
'''

# Imports
from pathlib import Path
import requests
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Forever filters",
    page_icon="🚲",
    layout="wide",
)

#set up containers to work in 
header = st.container()

# title
with header:
    st.title('🚲 Bikes forever store')

# Selector file with 200mb max size
uploaded_xlxs_file = st.file_uploader(
        "Upload your xlxs file into s3 storage", type=["xlsx"],
            accept_multiple_files=False,
            help="To activate 'wide mode', go to the hamburger menu > Settings > turn on 'wide mode'"
)

# Make request only if file had been upload
if uploaded_xlxs_file:

    # open and extract xlxs cheetsheet
    xlsFile = openpyxl.load_workbook(uploaded_xlxs_file)
    sheetnames = xlsFile.sheetnames

    st.text("Look into upload file")
    file_container = st.expander("Check your uploaded files")
    # Import datasets from all excel sheets
    for name in sheetnames:
        file_container.write(pd.read_excel(uploaded_xlxs_file, sheet_name=name, skiprows=[0]))

    upload = st.button("Upload to s3")

    if upload:
        # Request to api endpoint

        response = requests.post(
            'http://127.0.0.1:8000/dataset/upload_s3',
            data={"type": "multipart/form-data"},
            files={"file": uploaded_xlxs_file})

        # transform json object into dictionary
        response = response.json()

        if response['saved'] is True:
            st.write('Uploaded to s3')
        else:
            st.write('Can´t upload')

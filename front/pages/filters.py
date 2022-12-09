'''To create filters to process datasets

Receives:
raw excel dataset

Returns:
cleaned and processed tables
'''

# Imports
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)
from functions import filter_dataframe
from PIL import Image
import requests
import openpyxl
import pandas as pd
import streamlit as st
import plotly.express as px

st.set_page_config(
    page_title="Forever filters",
    page_icon="🚲",
    layout="wide",
)

#set up containers to work in 
header = st.container()
upload_xlxs = st.container()
upload_csv = st.container()
upload = st.container()
dataset = st.container()
features = st.container()
opt = st.container()


uploaded_xlxs_file=""
uploaded_csv_file=""


# title
with header:
    st.title('🚲 Bikes forever store')
# ask the user to upload file
with upload:
    # Selector file with 200mb max size
    uploaded_xlxs_file = st.file_uploader(
            "Upload your xlxs file", type=["xlsx"],
                accept_multiple_files=False,  
                help="To activate 'wide mode', go to the hamburger menu > Settings > turn on 'wide mode'"
    )

    # Make request only if file had been upload
    if uploaded_xlxs_file:
        # Request to api endpoint
        response = requests.post(
            'http://127.0.0.1:8000/dataset/upload',
            data={"type": "multipart/form-data"},
            files={"file": uploaded_xlxs_file})

        # transform json object into dictionary
        response = response.json()

        # open and extract xlxs cheetsheet
        xlsFile = openpyxl.load_workbook(response['path'])
        sheetnames = xlsFile.sheetnames
        
        st.text("Look into upload file")
        file_container = st.expander("Check your uploaded files")
        # Import datasets from all excel sheets
        for name in sheetnames:
            file_container.write(pd.read_excel(uploaded_xlxs_file, sheet_name=name, skiprows=[0]))

            
        option = st.sidebar.selectbox('Select type of file to see', ("Processed file", "Raw file"))

        if option == "Processed file":
            with dataset:
                dats = requests.get('http://127.0.0.1:8000/dataset/get_data')

                # Convert json object into python dict
                dats = dats.json()
                file_names = []
                st.text("Look into modified Files")
                file_container = st.expander("Modified csv files")
                
                #clean file names discarding file path and extensions
                for i in dats:
                    file_container.write(pd.read_csv(dats[i]))
                    file_names.append(dats[i].split("/")[-1].split(".")[-2])        

                #displays filtered dataframe on screen
                with features:
                    st.write("Filter tables:")
                    df_option = st.selectbox('Select table',file_names)
                    st.write(df_option)
                
                    for index, name in enumerate(file_names):
                        if df_option == file_names[index]:
                            df_filtered = filter_dataframe(pd.read_csv(dats[name]))
                            st.dataframe(df_filtered)
                            st.download_button('Download file', data=df_filtered.to_csv(),file_name=f'{name}.csv')
                            
        elif option == "Raw file":
            with st.sidebar:
                name_option = st.selectbox('Select Table that you would like to see', sheetnames)
            if name_option in sheetnames:
                df_filtered = filter_dataframe(pd.read_excel(uploaded_xlxs_file, sheet_name=name_option,skiprows=[0]))
                st.dataframe(df_filtered)
                st.download_button('Download file', data=df_filtered.to_csv(),file_name=f'{uploaded_xlxs_file.name}_filtered.csv')


st.text('Fuente: https://www.kaggle.com/datasets/archit9406/customer-transaction-dataset')
